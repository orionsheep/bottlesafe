from __future__ import annotations

import argparse
import hashlib
import json
import random
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from PIL import Image


HAZMAT = {
    "01": ("toxic", "critical", "有毒危险品标志"),
    "02": ("oxidizer", "high", "氧化性气体标志"),
    "03": ("flammable", "high", "易燃气体标志"),
    "04": ("flammable", "high", "易燃固体标志"),
    "05": ("corrosive", "high", "腐蚀性物质标志"),
    "06": ("dangerous", "high", "危险品标志"),
    "07": ("pressurized", "medium", "非易燃压缩气体标志"),
    "08": ("oxidizer", "high", "有机过氧化物标志"),
    "09": ("explosive", "critical", "爆炸品标志"),
    "10": ("radioactive", "critical", "放射性物质标志"),
    "11": ("toxic_inhalation", "critical", "吸入危害标志"),
    "12": ("flammable", "high", "自燃物质标志"),
    "13": ("biohazard", "critical", "感染性物质标志"),
}

SELECTED = {
    "Kitchen, Glass & Drain": 35,
    "Disinfectant Spray & Cleaners": 35,
    "Floor & Other Cleaners": 35,
    "Toilet Cleaners": 35,
    "Metal, Furniture Cleaner": 25,
    "Imported Cleaners": 25,
    "Detergents & Dishwash": 45,
    "Freshners & Repellents": 35,
}


def rows_from_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(x) if x is not None else "" for x in next(rows)]
    result = []
    for values in rows:
        row = dict(zip(headers, values))
        if row.get("Image Link") and row.get("SKU Name"):
            result.append(row)
    wb.close()
    return result


def download_one(item: dict, image_dir: Path) -> tuple[dict, str | None]:
    url = str(item["Image Link"])
    ean = str(item.get("EAN Code") or hashlib.sha1(url.encode()).hexdigest()[:12]).replace("/", "_")
    target = image_dir / f"{ean}.jpg"
    if not target.exists():
        try:
            req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urlopen(req, timeout=30) as response, target.open("wb") as f:
                shutil.copyfileobj(response, f)
        except Exception as exc:
            target.unlink(missing_ok=True)
            return item, str(exc)
    try:
        with Image.open(target) as image:
            image.verify()
    except Exception as exc:
        target.unlink(missing_ok=True)
        return item, f"invalid image: {exc}"
    item["local_image"] = str(target.resolve())
    return item, None


def product_answer(row: dict) -> dict:
    category = str(row.get("Sub-sub-Category") or row.get("Sub-Category") or "清洁用品")
    name = str(row["SKU Name"]).strip()
    brand = str(row.get("Brand") or "").strip() or None
    barcode = str(row.get("EAN Code") or "").strip() or None
    return {
        "product": {"name": name, "brand": brand, "category": category, "barcode": barcode, "manufacturer": None},
        "visual_evidence": [f"公开商品数据将该图片标注为 {category} 类产品"],
        "hazards": [], "ingredients": [], "signal_words": [],
        "safe_storage": ["按照产品包装标签要求保存，并远离儿童和宠物"],
        "do_not_mix_with": [],
        "first_aid": {"ingestion": None, "inhalation": None, "eye_contact": None, "skin_contact": None},
        "uncertainties": ["公开商品正面图和商品元数据不足以确认完整成分、浓度及危险等级"],
        "needs_more_images": ["产品背面成分与警示标签", "条码近照"],
        "risk_level": "unknown",
        "summary": f"识别为 {brand + ' ' if brand else ''}{name}，类别为 {category}；需要背面标签才能进一步判断化学风险。",
    }


def hazmat_answer(code: str) -> dict:
    kind, level, label = HAZMAT[code]
    return {
        "product": {"name": None, "brand": None, "category": "危险品标志", "barcode": None, "manufacturer": None},
        "visual_evidence": [f"图片中可见{label}"],
        "hazards": [{"type": kind, "severity": level, "evidence": label, "confidence": 0.98}],
        "ingredients": [], "signal_words": [],
        "safe_storage": ["依据实际物质的安全数据表和包装标签进行隔离储存"],
        "do_not_mix_with": [],
        "first_aid": {"ingestion": None, "inhalation": None, "eye_contact": None, "skin_contact": None},
        "uncertainties": ["危险标志只能说明危害类别，无法单独确认具体物质和浓度"],
        "needs_more_images": ["完整容器标签", "物质名称或 UN 编号", "安全数据表信息"],
        "risk_level": level,
        "summary": f"检测到{label}，应按高风险物质处理并查阅包装标签及安全数据表。",
    }


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--bigbasket", type=Path, required=True)
    p.add_argument("--hazmat", type=Path, required=True)
    p.add_argument("--output", type=Path, required=True)
    p.add_argument("--seed", type=int, default=42)
    args = p.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)
    image_dir = args.output / "images"
    image_dir.mkdir(exist_ok=True)
    rng = random.Random(args.seed)

    selected: list[dict] = []
    for path in args.bigbasket.rglob("product_data.xlsx"):
        category = path.parent.name
        if category not in SELECTED:
            continue
        rows = rows_from_xlsx(path)
        rng.shuffle(rows)
        unique, seen = [], set()
        for row in rows:
            key = str(row.get("EAN Code") or row.get("Image Link"))
            if key not in seen:
                seen.add(key)
                unique.append(row)
        selected.extend(unique[: SELECTED[category]])

    downloaded, failures = [], []
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = [pool.submit(download_one, row, image_dir) for row in selected]
        for future in as_completed(futures):
            row, error = future.result()
            (failures if error else downloaded).append((row, error) if error else row)

    records = [{
        "image": row["local_image"],
        "group_id": f"bigbasket-{row.get('EAN Code') or hashlib.sha1(row['Image Link'].encode()).hexdigest()[:12]}",
        "source": "BigBasket Products Dataset (MIT)",
        "answer": product_answer(row),
    } for row in downloaded]

    distance_root = args.hazmat / "images" / "DifferentDistances"
    for image in sorted(distance_root.glob("*/*.jpg")):
        code = image.stem.zfill(2)
        if code in HAZMAT:
            records.append({
                "image": str(image.resolve()),
                "group_id": f"hazmat-{code}-{image.parent.name}",
                "source": "HAZMAT13 (MIT)",
                "answer": hazmat_answer(code),
            })

    rng.shuffle(records)
    with (args.output / "all.jsonl").open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    with (args.output / "download_failures.jsonl").open("w", encoding="utf-8") as f:
        for row, error in failures:
            f.write(json.dumps({"url": row.get("Image Link"), "error": error}, ensure_ascii=False) + "\n")
    print(json.dumps({"requested_products": len(selected), "downloaded_products": len(downloaded), "hazmat_images": len(records)-len(downloaded), "total": len(records), "failures": len(failures)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
